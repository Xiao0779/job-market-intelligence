#!/usr/bin/env python3
"""
Refresh a privacy-safe job search analytics snapshot for GitHub.

This script merges:
1. ~/Desktop/career.xlsx
2. ~/Desktop/career-ops/data/applications.md
3. ~/Desktop/career-ops/data/pipeline.md

Outputs:
- data/jobs.db (gitignored local database)
- data/daily_summary.md
- data/public_dashboard.json
- data/top_pipeline.csv
- data/fig*.png
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SQL_DIR = ROOT / "sql"
DB_PATH = DATA_DIR / "jobs.db"
SUMMARY_PATH = DATA_DIR / "daily_summary.md"
JSON_PATH = DATA_DIR / "public_dashboard.json"
PIPELINE_CSV_PATH = DATA_DIR / "top_pipeline.csv"
DESKTOP = Path.home() / "Desktop"
CAREER_XLSX = DESKTOP / "career.xlsx"
CAREER_OPS = DESKTOP / "career-ops"
TRACKER_MD = CAREER_OPS / "data" / "applications.md"
PIPELINE_MD = CAREER_OPS / "data" / "pipeline.md"
PROFILE_YML = CAREER_OPS / "config" / "profile.yml"
SCHEMA_SQL = SQL_DIR / "schema.sql"


STATUS_MAP = {
    "waiting": "applied",
    "applied": "applied",
    "evaluated": "evaluated",
    "responded": "responded",
    "interview": "interview",
    "offer": "offer",
    "rejected": "rejected",
    "discarded": "discarded",
    "skip": "skip",
    "sent": "sent",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().replace(microsecond=0).isoformat()


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_company_role(value: str) -> str:
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def normalize_status(value: str) -> str:
    raw = clean_text(value).lower()
    aliases = {
        "waiting": "applied",
        "applied": "applied",
        "evaluated": "evaluated",
        "responded": "responded",
        "interview": "interview",
        "offer": "offer",
        "rejected": "rejected",
        "discarded": "discarded",
        "skip": "skip",
        "no aplicar": "skip",
        "descartado": "discarded",
        "rechazado": "rejected",
        "aplicado": "applied",
        "respondido": "responded",
        "entrevista": "interview",
        "oferta": "offer",
        "-": "unknown",
        "": "unknown",
    }
    return aliases.get(raw, raw)


def parse_score(value: str) -> float | None:
    text = clean_text(value)
    if not text or text in {"-", "N/A"}:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return float(match.group(1)) if match else None


def parse_date(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if "00:00:00" in text:
        text = text.split(" ")[0]
    match = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    return match.group(1) if match else text


def looks_like_location(value: str) -> bool:
    text = clean_text(value)
    if not text:
        return False
    lowered = text.lower()
    bad_prefixes = ("linkedin:", "resume", "cover letter", "oa", "email")
    if lowered.startswith(bad_prefixes):
        return False
    location_tokens = [
        "remote",
        "san ",
        "new york",
        "los angeles",
        "mountain view",
        "redmond",
        "seattle",
        "beijing",
        "chicago",
        "atlanta",
        "washington",
        "burbank",
        "cupertino",
        "portland",
        "california",
        "francisco",
        "diego",
    ]
    return any(token in lowered for token in location_tokens)


def extract_report_path(value: str) -> str:
    match = re.search(r"\]\(([^)]+)\)", value)
    return clean_text(match.group(1)) if match else ""


def extract_location_from_notes(notes: str) -> str:
    parts = [clean_text(part) for part in notes.split("|")]
    for part in parts:
        if looks_like_location(part):
            return part
    return ""


@dataclass
class Record:
    company: str
    role: str
    location: str = ""
    date_found: str = ""
    date_applied: str = ""
    jd_url: str = ""
    raw_status: str = ""
    normalized_status: str = "unknown"
    score: float | None = None
    report_path: str = ""
    pdf_ready: int = 0
    in_pipeline: int = 0
    source_labels: set[str] = field(default_factory=set)
    source_detail: set[str] = field(default_factory=set)
    linkedin_contact: str = ""
    linkedin_status: str = ""
    oa_status: str = ""
    interview_status: str = ""
    notes: list[str] = field(default_factory=list)
    excel_present: int = 0
    career_ops_present: int = 0

    @property
    def external_key(self) -> str:
        base = [normalize_company_role(self.company), normalize_company_role(self.role)]
        location = normalize_company_role(self.location)
        if location:
            base.append(location)
        elif self.jd_url:
            base.append(normalize_company_role(self.jd_url))
        return "::".join(filter(None, base))

    @property
    def remote(self) -> int:
        return 1 if "remote" in self.location.lower() else 0


def parse_excel_rows() -> list[dict]:
    wb = load_workbook(CAREER_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    parsed = []
    for row in rows[1:]:
        values = [clean_text(cell) for cell in row[:10]]
        if not any(values):
            continue
        date_found, company, role, location, oa, interview, app_status, li_contact, li_status, jd_url = values
        if not (company or role or location):
            continue
        parsed.append(
            {
                "date_found": parse_date(date_found),
                "company": company,
                "role": role,
                "location": location,
                "oa_status": oa,
                "interview_status": interview,
                "raw_status": app_status,
                "normalized_status": normalize_status(app_status),
                "linkedin_contact": li_contact,
                "linkedin_status": li_status.lower() if li_status else "",
                "jd_url": jd_url,
            }
        )
    return parsed


def parse_career_ops_tracker() -> list[dict]:
    lines = TRACKER_MD.read_text(encoding="utf-8").splitlines()
    rows = []
    for line in lines:
        if not line.startswith("|"):
            continue
        parts = [clean_text(part) for part in line.split("|")[1:-1]]
        if len(parts) < 8 or parts[0] in {"#", "---"}:
            continue
        try:
            int(parts[0])
        except ValueError:
            continue
        notes = " | ".join(parts[8:]).strip()
        rows.append(
            {
                "date_found": parse_date(parts[1]),
                "company": parts[2],
                "role": parts[3],
                "location": extract_location_from_notes(notes),
                "score": parse_score(parts[4]),
                "raw_status": parts[5],
                "normalized_status": normalize_status(parts[5]),
                "pdf_ready": 1 if "✅" in parts[6] else 0,
                "report_path": extract_report_path(parts[7]),
                "notes": notes,
            }
        )
    return rows


def parse_pipeline_rows() -> list[dict]:
    rows = []
    for line in PIPELINE_MD.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped.startswith("- [ ] "):
            continue
        payload = stripped[6:]
        parts = [clean_text(part) for part in payload.split("|")]
        if len(parts) < 3:
            continue
        url, company, role = parts[:3]
        rows.append({"jd_url": url, "company": company, "role": role})
    return rows


def choose_key(index: dict[str, Record], company: str, role: str, location: str, jd_url: str) -> str:
    exact = Record(company=company, role=role, location=location, jd_url=jd_url).external_key
    if exact in index:
        return exact
    base_prefix = "::".join([normalize_company_role(company), normalize_company_role(role)])
    candidates = [key for key in index if key.startswith(base_prefix)]
    if len(candidates) == 1:
        return candidates[0]
    return exact


def merge_sources(excel_rows: Iterable[dict], tracker_rows: Iterable[dict], pipeline_rows: Iterable[dict]) -> dict[str, Record]:
    index: dict[str, Record] = {}

    for row in excel_rows:
        key = choose_key(index, row["company"], row["role"], row["location"], row["jd_url"])
        record = index.setdefault(key, Record(company=row["company"], role=row["role"], location=row["location"]))
        record.location = record.location or row["location"]
        record.date_found = record.date_found or row["date_found"]
        record.date_applied = record.date_applied or row["date_found"]
        record.jd_url = record.jd_url or row["jd_url"]
        record.raw_status = record.raw_status or row["raw_status"]
        record.normalized_status = row["normalized_status"] if row["normalized_status"] != "unknown" else record.normalized_status
        record.linkedin_contact = record.linkedin_contact or row["linkedin_contact"]
        record.linkedin_status = record.linkedin_status or row["linkedin_status"]
        record.oa_status = record.oa_status or row["oa_status"]
        record.interview_status = record.interview_status or row["interview_status"]
        record.source_labels.add("career.xlsx")
        record.source_detail.add("excel")
        record.excel_present = 1

    for row in tracker_rows:
        key = choose_key(index, row["company"], row["role"], row["location"], "")
        record = index.setdefault(key, Record(company=row["company"], role=row["role"], location=row["location"]))
        record.location = record.location or row["location"]
        record.date_found = record.date_found or row["date_found"]
        record.date_applied = record.date_applied or row["date_found"]
        record.raw_status = row["raw_status"] or record.raw_status
        if row["normalized_status"] != "unknown":
            record.normalized_status = row["normalized_status"]
        if row["score"] is not None:
            record.score = row["score"]
        record.report_path = row["report_path"] or record.report_path
        record.pdf_ready = max(record.pdf_ready, row["pdf_ready"])
        if row["notes"]:
            record.notes.append(row["notes"])
        record.source_labels.add("career-ops")
        record.source_detail.add("tracker")
        record.career_ops_present = 1

    for row in pipeline_rows:
        key = choose_key(index, row["company"], row["role"], "", row["jd_url"])
        record = index.setdefault(key, Record(company=row["company"], role=row["role"], jd_url=row["jd_url"]))
        record.jd_url = record.jd_url or row["jd_url"]
        record.in_pipeline = 1
        record.source_labels.add("career-ops")
        record.source_detail.add("pipeline")

    return index


def load_profile() -> dict:
    if not PROFILE_YML.exists():
        return {}
    with PROFILE_YML.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def init_db(records: dict[str, Record]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)
    with SCHEMA_SQL.open("r", encoding="utf-8") as fh:
        conn.executescript(fh.read())

    now = now_iso()
    for record in records.values():
        source = ", ".join(sorted(record.source_labels))
        source_detail = ", ".join(sorted(record.source_detail))
        notes = " | ".join(dict.fromkeys([clean_text(note) for note in record.notes if clean_text(note)]))
        cur = conn.execute(
            """
            INSERT INTO jobs (
                external_key, date_found, company, role, location, remote, jd_url,
                source, source_detail, score, report_path, pdf_ready, in_pipeline,
                raw_status, normalized_status, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record.external_key,
                record.date_found,
                record.company,
                record.role,
                record.location,
                record.remote,
                record.jd_url,
                source,
                source_detail,
                record.score,
                record.report_path,
                record.pdf_ready,
                record.in_pipeline,
                record.raw_status,
                record.normalized_status,
                now,
            ),
        )
        conn.execute(
            """
            INSERT INTO applications (
                job_id, date_applied, app_status, source_status, linkedin_contact,
                linkedin_status, oa_status, interview_status, notes, excel_present,
                career_ops_present
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cur.lastrowid,
                record.date_applied or record.date_found,
                record.normalized_status,
                record.raw_status,
                record.linkedin_contact,
                record.linkedin_status,
                record.oa_status,
                record.interview_status,
                notes,
                record.excel_present,
                record.career_ops_present,
            ),
        )

    conn.commit()
    conn.close()


def generate_charts(records: dict[str, Record]) -> None:
    status_counts = Counter(record.normalized_status for record in records.values() if record.normalized_status != "unknown")
    company_counts = Counter(record.company for record in records.values())
    weekly_counts = Counter(record.date_found[:7] for record in records.values() if record.date_found)
    location_counts = Counter(record.location or "Unknown" for record in records.values() if record.location)

    fig1 = DATA_DIR / "fig1_status_breakdown.png"
    fig2 = DATA_DIR / "fig2_weekly_cadence.png"
    fig3 = DATA_DIR / "fig3_top_companies.png"
    fig4 = DATA_DIR / "fig4_location_distribution.png"

    plt.style.use("seaborn-v0_8-whitegrid")

    plt.figure(figsize=(7, 4))
    labels = list(status_counts.keys())
    values = [status_counts[label] for label in labels]
    plt.bar(labels, values, color="#2563eb")
    plt.title("Application Status Breakdown")
    plt.xticks(rotation=25, ha="right")
    plt.tight_layout()
    plt.savefig(fig1, dpi=160)
    plt.close()

    plt.figure(figsize=(7, 4))
    week_labels = sorted(weekly_counts.keys())
    week_values = [weekly_counts[label] for label in week_labels]
    plt.plot(week_labels, week_values, marker="o", color="#0f766e")
    plt.title("Monthly Application Cadence")
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()
    plt.savefig(fig2, dpi=160)
    plt.close()

    plt.figure(figsize=(7, 4))
    top_companies = company_counts.most_common(10)
    plt.barh([item[0] for item in top_companies][::-1], [item[1] for item in top_companies][::-1], color="#7c3aed")
    plt.title("Top Companies by Tracking Volume")
    plt.tight_layout()
    plt.savefig(fig3, dpi=160)
    plt.close()

    plt.figure(figsize=(7, 4))
    top_locations = location_counts.most_common(10)
    plt.barh([item[0] for item in top_locations][::-1], [item[1] for item in top_locations][::-1], color="#ea580c")
    plt.title("Top Locations")
    plt.tight_layout()
    plt.savefig(fig4, dpi=160)
    plt.close()


def write_public_outputs(records: dict[str, Record], profile: dict) -> None:
    refreshed_at = now_iso()
    status_counts = Counter(record.normalized_status for record in records.values() if record.normalized_status != "unknown")
    total = len(records)
    active = sum(status_counts.get(key, 0) for key in ("applied", "responded", "interview", "evaluated"))
    submitted = sum(status_counts.get(key, 0) for key in ("applied", "responded", "interview", "offer", "rejected"))
    positive = sum(status_counts.get(key, 0) for key in ("responded", "interview", "offer"))
    response_rate = round((positive / submitted) * 100, 1) if submitted else 0.0
    linkedin_sent = sum(1 for record in records.values() if record.linkedin_status == "sent")
    scored = [record.score for record in records.values() if record.score is not None]
    avg_score = round(sum(scored) / len(scored), 2) if scored else None
    pending_pipeline = [record for record in records.values() if record.in_pipeline]
    top_companies = Counter(record.company for record in records.values()).most_common(10)
    recent = sorted(
        [record for record in records.values() if record.date_found],
        key=lambda item: item.date_found,
        reverse=True,
    )[:10]

    overdue = []
    today = datetime.now().date()
    for record in records.values():
        if record.normalized_status != "applied" or not record.date_applied:
            continue
        applied_at = datetime.fromisoformat(record.date_applied).date()
        age = (today - applied_at).days
        if age >= 14:
            overdue.append((age, record))
    overdue.sort(reverse=True, key=lambda item: item[0])

    primary_roles = profile.get("target_roles", {}).get("primary", [])
    secondary_roles = profile.get("target_roles", {}).get("secondary", [])
    headline = profile.get("narrative", {}).get("headline", "")
    target_keywords = [
        "data analyst",
        "data operations",
        "data scientist",
        "business intelligence",
        "business analyst",
        "operations analyst",
        "commercial ops",
        "product analyst",
        "research analyst",
        "analytics engineer",
        "analytics developer",
        "machine learning engineer",
        "ml engineer",
        "nlp engineer",
        "applied scientist",
        "ai research engineer",
        "strategy and analytics",
    ]
    seniority_blockers = [
        "senior",
        "sr.",
        " sr ",
        "lead",
        "staff",
        "principal",
        "manager",
        "director",
        "head",
    ]
    non_us_markers = [
        "canada",
        "ontario",
        "toronto",
        "london",
        "uk",
        "japan",
        "tokyo",
        "india",
        "singapore",
        "korea",
        "qatar",
        "germany",
        "berlin",
        "mexico",
        "dublin",
    ]

    def is_recommended_pipeline(record: Record) -> bool:
        role_lower = record.role.lower()
        location_lower = record.location.lower()
        if not any(keyword in role_lower for keyword in target_keywords):
            return False
        if any(blocker in role_lower for blocker in seniority_blockers):
            if not any(token in role_lower for token in ("new grad", "entry", "associate", "junior")):
                return False
        if any(marker in location_lower for marker in non_us_markers):
            return False
        return True

    recommended_pipeline = [record for record in pending_pipeline if is_recommended_pipeline(record)]

    with PIPELINE_CSV_PATH.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["company", "role", "location", "jd_url"])
        for record in sorted(recommended_pipeline, key=lambda item: (item.company, item.role)):
            writer.writerow([record.company, record.role, record.location, record.jd_url])

    public_json = {
        "refreshed_at": refreshed_at,
        "headline": headline,
        "primary_roles": primary_roles,
        "secondary_roles": secondary_roles,
        "metrics": {
            "tracked_roles": total,
            "active_roles": active,
            "response_rate_pct": response_rate,
            "linkedin_sent": linkedin_sent,
            "pending_pipeline": len(pending_pipeline),
            "recommended_pipeline": len(recommended_pipeline),
            "avg_evaluated_score": avg_score,
        },
        "status_counts": dict(status_counts),
        "top_companies": [{"company": company, "count": count} for company, count in top_companies],
        "recent_roles": [
            {
                "date_found": record.date_found,
                "company": record.company,
                "role": record.role,
                "location": record.location,
                "status": record.normalized_status,
            }
            for record in recent
        ],
    }
    JSON_PATH.write_text(json.dumps(public_json, indent=2), encoding="utf-8")

    lines = [
        "# Daily Job Search Summary",
        "",
        f"_Last refreshed: {refreshed_at}_",
        "",
        "## Positioning",
        "",
        f"- Headline: {headline or 'Not set'}",
        f"- Primary targets: {', '.join(primary_roles) if primary_roles else 'Not set'}",
        f"- Secondary targets: {', '.join(secondary_roles) if secondary_roles else 'Not set'}",
        "",
        "## Snapshot",
        "",
        f"- Tracked roles: {total}",
        f"- Active roles: {active}",
        f"- Response rate: {response_rate}%",
        f"- LinkedIn outreach sent: {linkedin_sent}",
        f"- Pending pipeline items: {len(pending_pipeline)}",
        f"- Recommended pipeline items: {len(recommended_pipeline)}",
        f"- Average evaluated score: {avg_score if avg_score is not None else 'N/A'}",
        "",
        "## Status Breakdown",
        "",
    ]

    for status, count in sorted(status_counts.items()):
        lines.append(f"- {status}: {count}")

    lines.extend(
        [
            "",
            "## Recent Activity",
            "",
        ]
    )
    for record in recent:
        lines.append(f"- {record.date_found}: {record.company} | {record.role} | {record.normalized_status}")

    lines.extend(
        [
            "",
            "## Recommended Opportunity Queue",
            "",
        ]
    )
    if recommended_pipeline:
        for record in sorted(recommended_pipeline, key=lambda item: (item.company, item.role))[:15]:
            suffix = f" | {record.jd_url}" if record.jd_url else ""
            lines.append(f"- {record.company} | {record.role}{suffix}")
    else:
        lines.append("- No recommended pipeline items right now.")

    lines.extend(
        [
            "",
            "## Needs Attention",
            "",
        ]
    )
    if overdue:
        for age, record in overdue[:10]:
            lines.append(f"- {record.company} | {record.role} | {age} days with no response")
    else:
        lines.append("- No overdue follow-ups in the public snapshot.")

    SUMMARY_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not CAREER_XLSX.exists():
        raise FileNotFoundError(f"Missing source file: {CAREER_XLSX}")
    if not TRACKER_MD.exists():
        raise FileNotFoundError(f"Missing source file: {TRACKER_MD}")
    if not PIPELINE_MD.exists():
        raise FileNotFoundError(f"Missing source file: {PIPELINE_MD}")

    excel_rows = parse_excel_rows()
    tracker_rows = parse_career_ops_tracker()
    pipeline_rows = parse_pipeline_rows()
    profile = load_profile()
    records = merge_sources(excel_rows, tracker_rows, pipeline_rows)

    init_db(records)
    generate_charts(records)
    write_public_outputs(records, profile)

    print(f"Refreshed analytics for {len(records)} merged roles.")
    print(f"Wrote database: {DB_PATH}")
    print(f"Wrote summary: {SUMMARY_PATH}")
    print(f"Wrote dashboard JSON: {JSON_PATH}")


if __name__ == "__main__":
    main()
